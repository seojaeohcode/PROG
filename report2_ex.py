import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotInteractableException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from difflib import SequenceMatcher
import time

def parse_table(page_source):
    """
    HTML에서 table.esg_lank_tbl -> td[name="com_nm"]의 회사명을 추출
    """
    soup = BeautifulSoup(page_source, "html.parser")
    table = soup.select_one("table.esg_lank_tbl")
    if not table:
        return []
    rows = table.select("tbody tr")
    results = []
    for row in rows:
        td_company = row.select_one('td[name="com_nm"]')
        if td_company:
            link_tag = td_company.select_one("a")
            comp_name = link_tag.get_text(strip=True) if link_tag else td_company.get_text(strip=True)
            results.append(comp_name)
    return results

def get_governance_report_companies():
    webdriver_options = webdriver.ChromeOptions()
    webdriver_options.add_argument("--no-sandbox")
    webdriver_options.add_argument("--disable-dev-shm-usage")
    webdriver_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=webdriver_options)

    try:
        url = "https://esg.krx.co.kr/contents/02/02040000/ESG02040000.jsp"
        driver.get(url)
        time.sleep(3)

        from_date = driver.find_element(By.NAME, "fr_work_dt")
        to_date = driver.find_element(By.NAME, "to_work_dt")

        from_date.clear()
        from_date.send_keys("20240101")
        to_date.clear()
        to_date.send_keys("20241231")

        search_btn = driver.find_element(By.CSS_SELECTOR, "button.sch_btn")
        search_btn.click()
        time.sleep(5)

        all_companies = []
        page_num = 1

        while True:
            html = driver.page_source
            companies = parse_table(html)
            all_companies.extend(companies)

            try:
                next_li = driver.find_element(By.CSS_SELECTOR, "li.next:not(.disabled)")
            except NoSuchElementException:
                break

            next_a = next_li.find_element(By.TAG_NAME, "a")
            try:
                next_a.click()
                page_num += 1
                time.sleep(4)
            except ElementNotInteractableException:
                break

    finally:
        driver.quit()
    return set(all_companies)

def is_similar(name1, name2, threshold=0.8):
    """두 회사명 간 유사도 검사 (정확 일치 후 유사도 검사)"""
    if name1 == name2:
        return True
    similarity = SequenceMatcher(None, name1, name2).ratio()
    return similarity >= threshold

def update_governance_report_column(excel_file="company.xlsx"):
    governance_companies = get_governance_report_companies()
    wb = load_workbook(excel_file)
    ws = wb.active

    # E/F 열에 기업지배구조보고서 O/X 표시
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2), start=2):
        company_name = row[0].value
        ws.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=6)
        match_found = False

        for reported_company in governance_companies:
            if is_similar(company_name, reported_company, threshold=0.9):
                ws.cell(row=idx, column=5, value="O")
                match_found = True
                break

        if not match_found:
            ws.cell(row=idx, column=5, value="X")

    wb.save(excel_file)
    print(f"Excel 파일({excel_file})에 지배구조경영보고서 여부 업데이트 완료!")

if __name__ == "__main__":
    update_governance_report_column()
