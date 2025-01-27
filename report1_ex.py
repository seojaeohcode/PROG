import FinanceDataReader as fdr
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotInteractableException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import time
from bs4 import BeautifulSoup

webdriver_options = webdriver.ChromeOptions()
webdriver_options.add_argument('--no-sandbox')
webdriver_options.add_argument('--disable-dev-shm-usage')
webdriver_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/97.0.4692.71 Safari/537.36')
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=webdriver_options)

def parse_table(page_source):
    soup = BeautifulSoup(page_source, 'html.parser')
    table = soup.select_one("table.esg_lank_tbl")
    if not table:
        return []
    rows = table.select("tbody tr")
    companies = []
    for row in rows:
        td_company = row.select_one('td[name="com_abbrv"]')
        if td_company:
            raw_text = td_company.find(text=True, recursive=False)
            company_name = raw_text.strip() if raw_text else ""
            companies.append(company_name)
    return companies

def get_sustainability_report_companies():
    try:
        driver.get('https://esg.krx.co.kr/contents/02/02030000/ESG02030000.jsp')
        time.sleep(5)
        all_companies = []
        page_num = 1

        while True:
            page_source = driver.page_source
            companies = parse_table(page_source)
            all_companies.extend(companies)

            try:
                next_li = driver.find_element(By.CSS_SELECTOR, "li.next:not(.disabled)")
            except NoSuchElementException:
                break

            next_btn = next_li.find_element(By.TAG_NAME, "a")
            try:
                next_btn.click()
                page_num += 1
                time.sleep(5)
            except ElementNotInteractableException:
                break

    finally:
        driver.quit()
    return set(all_companies)

def is_subsidiary_reported(company, sustainability_companies):
    for reported_company in sustainability_companies:
        if reported_company in company or company in reported_company:
            return True
    return False

if __name__ == "__main__":
    sustainability_companies = get_sustainability_report_companies()
    excel_file = "company.xlsx"
    wb = load_workbook(excel_file)
    ws = wb.active

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2), start=2):
        company_name = row[0].value
        ws.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=4)
        if is_subsidiary_reported(company_name, sustainability_companies):
            ws.cell(row=idx, column=3, value="O")
        else:
            ws.cell(row=idx, column=3, value="X")

    wb.save(excel_file)
    print(f"Excel 파일({excel_file})에 지속가능경영보고서 여부 업데이트 완료!")
