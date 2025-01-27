import pandas as pd
from openpyxl import load_workbook

# 파일 로드
company_file = 'company.xlsx'
data_file = 'data.xlsx'

# 데이터 로드
company_df = pd.read_excel(company_file)
data_df = pd.read_excel(data_file)

# 발행기관별 카운트 계산
bond_issuers_count = data_df['발행기관'].value_counts().reset_index()
bond_issuers_count.columns = ['발행기관', '채권발행횟수']

# 데이터 확인
print("채권 발행기관별 카운트 목록:")
print(bond_issuers_count.to_string(index=False))

# 데이터 매칭 함수 정의 (공백 제거 및 대소문자 무시)
def is_match(company_name, issuer_name):
    if isinstance(company_name, str) and isinstance(issuer_name, str):
        return company_name.strip().lower() == issuer_name.strip().lower()
    return False

# company.xlsx 파일 열기
wb = load_workbook(company_file)
sheet = wb.active

# D열에 채권 발행 횟수 추가
for row in range(2, sheet.max_row + 1):
    company_name = sheet[f'A{row}'].value  # A열에 회사명이 있다고 가정
    if company_name:
        match_found = False
        for index, row_data in bond_issuers_count.iterrows():
            issuer_name = row_data['발행기관']
            if is_match(company_name, issuer_name):
                sheet[f'D{row}'].value = row_data['채권발행횟수']
                match_found = True
                break
        if not match_found:
            sheet[f'D{row}'].value = 0  # 매칭되지 않은 경우 0으로 설정

# 결과 저장
wb.save(company_file)

print("데이터 매칭 및 D열 채권 발행 횟수 업데이트 완료.")
