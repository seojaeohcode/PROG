import FinanceDataReader as fdr
import pandas as pd
from openpyxl import load_workbook

def get_top_market_cap_kospi(top_n=100):
    """
    KOSPI 시가총액 상위 top_n 개 기업을 조회하여 DataFrame으로 반환
    """
    # KOSPI 종목 데이터 불러오기
    kospi = fdr.StockListing('KOSPI')
    
    # 시가총액 내림차순 정렬 및 상위 top_n 추출
    kospi_sorted = kospi.sort_values(by="Marcap", ascending=False).head(top_n)
    
    # 필요한 컬럼만 정리
    result_df = kospi_sorted[['Code', 'Name', 'Marcap']].copy()
    
    # 시가총액을 원 단위로 변환 (기본 제공은 원 단위)
    result_df['Marcap'] = result_df['Marcap'].astype(int)
    
    return result_df

if __name__ == "__main__":
    # KOSPI 시가총액 상위 100개 종목 출력
    result_df = get_top_market_cap_kospi(100)
    print("KOSPI 시가총액 상위 100 종목:")
    print(result_df)

    # -----------------------
    # Excel(company.xlsx) 작성/수정 부분
    # -----------------------
    excel_file = "company.xlsx"  # 같은 경로에 존재한다고 가정

    try:
        # (1) 엑셀 파일 불러오기
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        print(f"오류: '{excel_file}' 파일을 찾을 수 없습니다.")
        exit(1)
    except Exception as e:
        print(f"엑셀 파일 로드 오류: {e}")
        exit(1)

    # (2) 첫 번째 시트 혹은 특정 시트 선택
    ws = wb.active  # wb["시트이름"] 으로 특정 시트 선택 가능

    # (3) 헤더(1행) 구성 (각 항목은 2열씩 차지)
    ws.merge_cells("A1:B1")
    ws["A1"] = "회사명"

    ws.merge_cells("C1:D1")
    ws["C1"] = "지속가능경영보고서"

    ws.merge_cells("E1:F1")
    ws["E1"] = "기업지배구조보고서"

    ws.merge_cells("G1:H1")
    ws["G1"] = "채권발행이력수"

    # (4) 시가총액 순으로 회사명 작성 (A/B열 병합 후 회사명 입력)
    company_names = result_df["Name"].tolist()
    for idx, company in enumerate(company_names, start=2):
        # A/B 열 병합
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=2)
        # 병합된 셀에 회사명 기입
        ws.cell(row=idx, column=1, value=company)

    # (5) 작업 내용 저장
    wb.save(excel_file)

    print(f"Excel 파일({excel_file})에 회사명 목록 업로드 완료!")
